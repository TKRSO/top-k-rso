import sys
import os

current_dir = os.path.dirname(os.path.abspath(__file__))
src_dir = os.path.dirname(current_dir)
sys.path.append(src_dir)
sys.path.append(os.path.join(src_dir, 'alg'))


from collections import defaultdict
from typing import List, Dict, Tuple
import pandas as pd
from models import Place, SquareGrid 
from config import COMBO, NUM_CELLS, GAMMAS, DATASET_NAMES

from alg.baseline_iadu import load_dataset, plot_selected, iadu
from alg.extension_sampling import grid_sampling, grid_weighted_sampling
from alg.biased_sampling import biased_sampling

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

# Using the name you requested
EXPERIMENT_NAME = "base_VS_extension_VS_biased"
SHAPES = DATASET_NAMES

def run_experiment():
    """
    Runs a simple comparison:
    1. Base IAdU (Main Paper) - Run 1 time
    2. Grid Sampling (Extension) - Run 1 time
    3. Biased Sampling (Baseline) - Run 1 time
    
    Logs scores AND times to Excel.
    Plots one sample run.
    """
    log = defaultdict(list)

    pdf_path = f"{EXPERIMENT_NAME}_plots.pdf"
    pdf_pages = PdfPages(pdf_path)

    for (K, k) in COMBO:
        for g in GAMMAS:
            W = K / (g * k)
            print(f"Comparing Base_IAdU vs. Grid_Sampling vs. Biased_Sampling | K={K}, k={k}, Gs={NUM_CELLS}")

            for shape in SHAPES:
                print(f"  Shape={shape}")
                S: List[Place] = load_dataset(shape, K)
                
                if not S:
                    print(f"    Skipping shape {shape} for K={K}, no data loaded.")
                    continue

                # --- 1. Run Base IAdU (IADU) ---
                # API: 3 args (S, k, W), 6 return values
                print(f"    Running Base_IAdU...")
                R_base, score_base, base_pss_sum, base_psr_sum, prep_time_base, selection_time_base = iadu(S, k, W)
                
                # --- 2. Run Biased Sampling (Baseline) ---
                # API: 3 args (S, k, W), 5 return values
                print(f"    Running Biased_Sampling...")
                R_biased, score_biased, biased_pss_sum, biased_psr_sum, selection_time_biased = biased_sampling(S, k, W)

                for G in NUM_CELLS:
                    print(f"      Running Grid_Sampling for G={G}...")
                    
                    try:
                        grid = SquareGrid(S, G)
                    except ValueError:
                        print(f"        Skipping G={G}. Invalid data for grid.")
                        continue
                    
                    lenCL = len(grid.get_full_cells())

                    # --- 3. Run Grid Sampling (Extension) ---
                    # API: 4 args (S, k, W, G), 8 return values
                    R_grid_samp, score_grid_samp, grid_samp_pss_sum, grid_samp_psr_sum, prep_time_grid, selection_time_grid, _, cell_stats_run = grid_sampling(S, k, W, G)

                    # --- LOGGING (using meaningful names) ---
                    log_entry = {
                        "shape": shape,
                        "K": K,
                        "k": k,
                        "W": W,
                        "G": G,
                        "lenCL": lenCL, 
                        "K/(k*g)": f"K/(k * {g})",

                        "base_iadu_hpfr": score_base,
                        "base_iadu_pss_sum": base_pss_sum,
                        "base_iadu_psr_sum": base_psr_sum,
                        
                        "grid_sampling_hpfr": score_grid_samp,
                        "grid_sampling_pss_sum": grid_samp_pss_sum,
                        "grid_sampling_psr_sum": grid_samp_psr_sum,

                        "biased_hpfr": score_biased,
                        "biased_pss_sum": biased_pss_sum,
                        "biased_psr_sum": biased_psr_sum,
                        
                        # Using your requested names
                        "baseline_prep_time": prep_time_base,
                        "gridsampling_prep_time": prep_time_grid,
                        "biasedsampling_prep_time": 0.0, # Biased sampling has no prep time
                        
                        "baseline_sel_time": selection_time_base,
                        "gridsampling_sel_time": selection_time_grid,
                        "biasedsampling_sel_time": selection_time_biased,
                        
                        "baseline_x_time": prep_time_base + selection_time_base,
                        "gridsampling_x_time": prep_time_grid + selection_time_grid,
                        "biasedsampling_x_time": selection_time_biased, # Total time is just selection time
                    }
                    log[(K, k, g, G)].append(log_entry)

                    # --- Plotting ---
                    fig, axes = plt.subplots(1, 3, figsize=(21, 7))
                    fig.suptitle(f"Shape: {shape}  |  K={K}, k={k}  |  G={G} (Ax={grid.Ax}, Ay={grid.Ay})  |  lenCL={lenCL}", fontsize=16)
                    
                    plot_selected(S, R_base, f"Base IAdU\nHPFR: {score_base: .4f}", axes[0], grid=grid)
                    
                    title_grid = f"Grid Sampling (Extension)\nHPFR: {score_grid_samp: .4f}"
                    plot_selected(S, R_grid_samp, title_grid, axes[1], grid=grid, cell_stats=cell_stats_run)
                    
                    title_biased = f"Biased Sampling\nHPFR: {score_biased: .4f}"
                    plot_selected(S, R_biased, title_biased, axes[2], grid=grid)
                    
                    pdf_pages.savefig(fig)
                    plt.close(fig) 

    pdf_pages.close()
    print(f"\nSuccessfully saved plots to {pdf_path}")

    avg_log = compute_average_log(log)
    save_outputs(log=avg_log, xlsx_name=f"{EXPERIMENT_NAME}.xlsx")

    
def save_outputs(log: Dict, xlsx_name: str):
    """
    Saves a simplified Excel file with the direct comparison.
    (Updated to use your requested column names)
    """
    def smart_round(value):
        if value is None:
            return None
        if value == 0:
            return 0.0
        elif abs(value) >= 0.01:
            return round(value, 3)
        else:
            return float(f"{value:.5f}") if abs(value) >= 1e-5 else f"{value:.1e}"

    all_rows = []
    for row in log.values():
        for k, v in list(row.items()):
            if isinstance(v, float):
                row[k] = smart_round(v)
            if k == "lenCL" and isinstance(v, float):
                 row[k] = round(v, 1)
        all_rows.append(row)

    df = pd.DataFrame(all_rows)

    if "shape" in df.columns:
        df.drop(columns=["shape"], inplace=True)

    sort_cols = [c for c in ["K", "k", "K/(k*g)", "G"] if c in df.columns]
    if sort_cols:
        df.sort_values(by=sort_cols, ascending=[True] * len(sort_cols), inplace=True)

    setup_cols = ["K", "k", "W", "K/(k*g)", "G", "lenCL"]
    
    score_cols = [
        "base_iadu_hpfr",
        "base_iadu_pss_sum",
        "base_iadu_psr_sum",
        
        "grid_sampling_hpfr",
        "grid_sampling_pss_sum",
        "grid_sampling_psr_sum",
        
        "biased_hpfr",
        "biased_pss_sum",
        "biased_psr_sum",
    ]
    
    # Using your requested names
    prep_cols = [
        "baseline_prep_time",
        "gridsampling_prep_time",
        "biasedsampling_prep_time",
    ]
    
    select_cols = [
        "baseline_sel_time",
        "gridsampling_sel_time",
        "biasedsampling_sel_time",
    ]
    
    total_cols = [
        "baseline_x_time",
        "gridsampling_x_time",
        "biasedsampling_x_time",
    ]
    
    all_cols = setup_cols + score_cols + prep_cols + select_cols + total_cols

    for col in all_cols:
        if col not in df.columns:
            df[col] = None
    df = df[all_cols]

    df.to_excel(xlsx_name, index=False)

    wb = load_workbook(xlsx_name)
    ws = wb.active
    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    group_fills = {
        "setup": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "score": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "prep": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "select": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "total": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
    }

    def apply_group_style(cols, fill, border_after=False):
        for i, col_name in enumerate(cols):
            if col_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(col_name) + 1
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    cell.fill = fill
                    current_border = cell.border
                    if border_after and i == len(cols) - 1:
                        cell.border = Border(left=current_border.left, right=Side(style='thick'),
                                              top=current_border.top, bottom=current_border.bottom)
                    else:
                        cell.border = thin_border

    apply_group_style(setup_cols, group_fills["setup"], border_after=True)
    apply_group_style(score_cols, group_fills["score"], border_after=True)
    apply_group_style(prep_cols, group_fills["prep"], border_after=True)
    apply_group_style(select_cols, group_fills["select"], border_after=True)
    apply_group_style(total_cols, group_fills["total"], border_after=False)

    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_len + 2

    wb.save(xlsx_name)
    print(f"Results saved to {xlsx_name}")

def compute_average_log(
    log: Dict[Tuple[int, int, int, int], List[Dict]]
) -> Dict[Tuple[int, int, int, int], Dict]:
    """
    Averages results per (K, k, g, G) key, across all shapes.
    """
    avg_log: Dict[Tuple[int, int, int, int], Dict] = {}

    for key, rows in log.items():  # key = (K, k, g, G)
        if not rows:
            continue
        
        out = {
            "K": key[0],
            "k": key[1],
            "W": rows[0]["W"],
            "K/(k*g)": rows[0]["K/(k*g)"],
            "G": key[3],
        }
        
        all_fields = set().union(*[r.keys() for r in rows])
        for fname in all_fields:
            if fname in {"shape", "K", "k", "g", "G", "W", "K/(k*g)"}:
                continue
            
            vals = [r[fname] for r in rows if isinstance(r.get(fname), (int, float))]
            if vals:
                out[fname] = sum(vals) / len(vals)
            elif fname not in out:
                out[fname] = None 

        avg_log[key] = out

    return avg_log
                        
if __name__ == "__main__":
    run_experiment()