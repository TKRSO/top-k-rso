import pandas as pd
import re
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExperimentLogger:
    def __init__(self, experiment_name, baseline_name="base_iadu", aggregate_datasets=True):
        self.filename = f"{experiment_name}.xlsx"
        self.baseline_name = baseline_name
        self.aggregate_datasets = aggregate_datasets
        self.logs = []
        
        # Metadata columns
        self.meta_columns = ["K", "k", "W", "wrf", "g*K/k", "G", "shape"]

    def log(self, row_dict):
        """ Logs a single result row. """
        self.logs.append(row_dict)

    def _calculate_derived_metrics(self, df):
        """ Calculates Diff% against Baseline (Adaptive) AND Biased vs Stratified. """
        # Primary Baseline Columns
        base_hpfr = f"{self.baseline_name}_hpfr"
        base_time = f"{self.baseline_name}_x_time"
        
        # Secondary Baseline (No R) Columns
        sec_base_name = "base_iadu_no_r"
        sec_base_hpfr = f"{sec_base_name}_hpfr"
        sec_base_time = f"{sec_base_name}_x_time"
        
        # 1. Identify all unique method prefixes (anything ending in _hpfr)
        all_cols = df.columns
        methods = {c.replace("_hpfr", "") for c in all_cols if c.endswith("_hpfr") and self.baseline_name not in c}
        
        for method in methods:
            # --- A. HPFR Diff % (Adaptive Baseline) ---
            score_col = f"{method}_hpfr"
            rf_col = f"{method}_rf_sum"
            diff_col = f"{method}_hpfr_diff%"
            
            if score_col in df.columns:
                def calc_adaptive_diff(row):
                    # 1. Determine Baseline Value
                    # Default: Primary Baseline
                    baseline_val = row.get(base_hpfr, 0)
                    
                    # Adaptive Logic: If method has NO rf_sum, use base_iadu_no_r
                    method_rf = row.get(rf_col, 0)
                    if method_rf == 0:
                        # Check if the secondary baseline exists in this row
                        if sec_base_hpfr in row and pd.notna(row[sec_base_hpfr]):
                            baseline_val = row[sec_base_hpfr]

                    # 2. Calculate Diff
                    val = row[score_col]
                    
                    if pd.isna(val) or pd.isna(baseline_val) or baseline_val == 0:
                        return 0
                    return (val - baseline_val) / baseline_val

                df[diff_col] = df.apply(calc_adaptive_diff, axis=1)

            # --- B. Speedup (Adaptive Baseline) ---
            # Calculates speedup relative to the SAME baseline used for accuracy
            time_col = f"{method}_x_time"
            speedup_col = f"{method}_speedup"
            
            if time_col in df.columns:
                def calc_adaptive_speedup(row):
                    # 1. Determine Baseline Time
                    base_t = row.get(base_time, 0)
                    
                    # Adaptive Logic: If method has NO rf_sum, use base_iadu_no_r time
                    method_rf = row.get(rf_col, 0)
                    if method_rf == 0:
                        if sec_base_time in row and pd.notna(row[sec_base_time]):
                            base_t = row[sec_base_time]

                    # 2. Calculate Speedup
                    t = row[time_col]
                    if pd.isna(t) or pd.isna(base_t) or t == 0:
                        return 0
                    return base_t / t

                df[speedup_col] = df.apply(calc_adaptive_speedup, axis=1)

        # 2. Custom: Biased vs Stratified Comparison
        biased_col = "biased_sampling_hpfr"
        strat_col = "stratified_sampling_hpfr"
        custom_diff_col = "biased_vs_stratified_diff%"

        if biased_col in df.columns and strat_col in df.columns:
            df[custom_diff_col] = df.apply(
                lambda row: (row[biased_col] - row[strat_col]) / row[strat_col] if row[strat_col] != 0 else 0,
                axis=1
            )
            
        return df

    def _organize_columns(self, df):
        """ Standardizes column order with strict blocks for every method found. """
        cols = list(df.columns)
        
        # 1. Start with Metadata
        final_cols = [c for c in self.meta_columns if c in cols]
        
        # 2. Baseline Block (Primary)
        base_block = [
            f"{self.baseline_name}_hpfr",
            f"{self.baseline_name}_rf_sum", 
            f"{self.baseline_name}_pss_sum", 
            f"{self.baseline_name}_psr_sum"
        ]
        final_cols.extend([c for c in base_block if c in cols])
        
        # 3. Dynamic Method Blocks
        # (This will automatically include base_iadu_no_r if it was run as a method)
        method_prefixes = sorted({c.replace("_hpfr", "") for c in cols if c.endswith("_hpfr") and self.baseline_name not in c})
        
        for method in method_prefixes:
            block = [
                f"{method}_hpfr", 
                f"{method}_hpfr_diff%", 
                f"{method}_rf_sum", 
                f"{method}_pss_sum", 
                f"{method}_psr_sum"
            ]
            final_cols.extend([c for c in block if c in cols])

        # 4. Custom Comparison Column
        if "biased_vs_stratified_diff%" in cols:
            final_cols.append("biased_vs_stratified_diff%")

        # 5. Times & Speedups (at the end)
        final_cols.extend(sorted([c for c in cols if "time" in c or "speedup" in c]))
        
        # 6. Catch-all
        existing = set(final_cols)
        remaining = [c for c in cols if c not in existing]
        final_cols.extend(remaining)
        
        return df[final_cols]

    def save(self):
        # ... (Rest of the save method remains exactly the same as your original) ...
        if not self.logs:
            print("No logs to save.")
            return

        df_raw = pd.DataFrame(self.logs)
        setting_keys = ["K", "k", "W", "wrf", "g*K/k", "G"]
        if not self.aggregate_datasets:
            setting_keys.append("shape")
        setting_keys = [k for k in setting_keys if k in df_raw.columns]
        agg_keys = setting_keys + (["run_id"] if "run_id" in df_raw.columns else [])

        df_proc = self._calculate_derived_metrics(df_raw)

        sort_keys = agg_keys.copy()
        if "shape" in sort_keys:
            sort_keys.remove("shape")
            sort_keys.insert(0, "shape")
        
        df_proc = df_proc.sort_values(by=sort_keys)

        final_rows = []
        prev_shape = None

        for name, group in df_proc.groupby(setting_keys, sort=False):
            if isinstance(name, tuple):
                current_map = dict(zip(setting_keys, name))
            else:
                current_map = {setting_keys[0]: name}
            
            def restore_meta(row_dict):
                for k, v in current_map.items():
                    row_dict[k] = v
                return row_dict

            current_shape = current_map.get("shape")
            if prev_shape is not None and current_shape != prev_shape:
                nq_row = {col: None for col in group.columns}
                nq_row["_row_type"] = "NEXT QUERY"
                final_rows.append(pd.DataFrame([nq_row]))
            prev_shape = current_shape

            if self.aggregate_datasets:
                main_row = group.mean(numeric_only=True).to_dict()
                final_rows.append(pd.DataFrame([restore_meta(main_row)]))
            else:
                if len(group) == 1:
                    final_rows.append(group)
            
            if len(group) > 1:
                target_cols = [c for c in group.columns if "diff%" in c or "speedup" in c]
                mean_row = group.mean(numeric_only=True).to_dict()
                mean_row["_row_type"] = "MEAN"
                final_rows.append(pd.DataFrame([restore_meta(mean_row)]))
                
                max_row = group.max(numeric_only=True)[target_cols].to_dict()
                max_row["_row_type"] = "MAX"
                final_rows.append(pd.DataFrame([restore_meta(max_row)]))
                
                min_row = group.min(numeric_only=True)[target_cols].to_dict()
                min_row["_row_type"] = "MIN"
                final_rows.append(pd.DataFrame([restore_meta(min_row)]))

                std_row = group.std(numeric_only=True)[target_cols].to_dict()
                std_row["_row_type"] = "STD"
                final_rows.append(pd.DataFrame([restore_meta(std_row)]))

            sep_row = {col: None for col in group.columns}
            sep_row["_row_type"] = "SEP"
            final_rows.append(pd.DataFrame([sep_row]))

        avg_key = "g*K/k" if "g*K/k" in df_proc.columns else ("W" if "W" in df_proc.columns else None)
        if avg_key:
            final_rows.append(pd.DataFrame([{"_row_type": "SEP"}]))
            w_groups = df_proc.groupby(avg_key)
            for w_val, w_group in w_groups:
                base_dict = {m: None for m in self.meta_columns}
                base_dict[avg_key] = w_val
                if "W" in self.meta_columns and avg_key != "W":
                     base_dict["W"] = w_group["W"].iloc[0] if "W" in w_group.columns else None

                avg_row = w_group.mean(numeric_only=True).to_dict()
                avg_row.update(base_dict)
                avg_row["_row_type"] = "W-AVG"
                final_rows.append(pd.DataFrame([avg_row]))
                
                # ... (rest of stats) ...
                max_row = w_group.max(numeric_only=True).to_dict()
                max_row.update(base_dict)
                max_row["_row_type"] = "W-MAX"
                final_rows.append(pd.DataFrame([max_row]))

                min_row = w_group.min(numeric_only=True).to_dict()
                min_row.update(base_dict)
                min_row["_row_type"] = "W-MIN"
                final_rows.append(pd.DataFrame([min_row]))

                std_row = w_group.std(numeric_only=True).to_dict()
                std_row.update(base_dict)
                std_row["_row_type"] = "W-STD"
                final_rows.append(pd.DataFrame([std_row]))

                final_rows.append(pd.DataFrame([{"_row_type": "SEP"}]))

        df_final = pd.concat(final_rows, ignore_index=True)
        df_final = self._organize_columns(df_final)

        if "_row_type" in df_final.columns:
            df_final.insert(0, "Type", df_final["_row_type"].fillna(""))
            df_final.drop(columns=["_row_type"], inplace=True)
        else:
            df_final.insert(0, "Type", "")

        cols_to_drop = ["run_id", "lenCL", "shape"] 
        df_final.drop(columns=[c for c in cols_to_drop if c in df_final.columns], inplace=True)

        diff_cols = [c for c in df_final.columns if "diff%" in c or "time" in c]
        meta_cols_present = [c for c in self.meta_columns if c in df_final.columns]
        diff_sheet_cols = ["Type"] + meta_cols_present + diff_cols
        df_diff = df_final[diff_sheet_cols].copy()

        try:
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                if not df_final.empty:
                    df_final_clean = df_final.copy()
                    df_final_clean["Type"] = df_final_clean["Type"].replace("SEP", "")
                    df_final_clean.to_excel(writer, sheet_name='Summary', index=False)
                
                if not df_diff.empty:
                    df_diff_clean = df_diff.copy()
                    df_diff_clean["Type"] = df_diff_clean["Type"].replace("SEP", "")
                    df_diff_clean.to_excel(writer, sheet_name='Diffs', index=False)
            
            self._apply_pro_styling()
            print(f"✔ Results saved to '{self.filename}'.")
        except Exception as e:
            print(f"❌ ERROR Saving Log: {e}")

    # _apply_pro_styling remains unchanged
    def _apply_pro_styling(self):
        try:
            wb = load_workbook(self.filename)
            c_header = "404040"
            c_meta   = "E7E6E6"
            method_palette = ["E2EFDA", "FBE5D6", "DDEBF7", "FFF2CC", "E2F0D9", "D9D9D9"]
            suffixes = ["_hpfr", "_rf_sum", "_pss_sum", "_psr_sum", "_x_time", "_speedup", "_diff%"]
            
            def get_method_name(header_val):
                if not header_val: return "meta"
                h = str(header_val)
                if h == "Type" or h in self.meta_columns: return "meta"
                if "biased_vs_stratified" in h: return "comparison"
                root = h
                for s in suffixes:
                    if root.endswith(s):
                        root = root.replace(s, "")
                        break
                return root

            for ws in wb.worksheets:
                col_map = {cell.value: i for i, cell in enumerate(ws[1], 1)}
                type_idx = col_map.get("Type")
                ws.freeze_panes = "C2" if type_idx else "B2"
                headers = [cell.value for cell in ws[1]]
                unique_methods = []
                for h in headers:
                    m = get_method_name(h)
                    if m not in unique_methods and m != "meta":
                        unique_methods.append(m)
                
                method_color_map = {}
                for idx, m_name in enumerate(unique_methods):
                    color = method_palette[idx % len(method_palette)]
                    method_color_map[m_name] = color
                method_color_map["meta"] = c_meta

                for i, col_cells in enumerate(ws.columns, 1):
                    header = str(col_cells[0].value)
                    col_letter = get_column_letter(i)
                    head_cell = col_cells[0]
                    head_cell.fill = PatternFill(start_color=c_header, fill_type="solid")
                    head_cell.font = Font(bold=True, color="FFFFFF")
                    head_cell.alignment = Alignment(horizontal="center", vertical="center")
                    valid_vals = [len(str(c.value)) for c in col_cells[:20] if c.value is not None]
                    max_len = max(valid_vals) if valid_vals else 0
                    ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 35)
                    m_name = get_method_name(header)
                    col_bg_hex = method_color_map.get(m_name, "FFFFFF") 
                    col_fill = PatternFill(start_color=col_bg_hex, fill_type="solid")
                    thin_border = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"))
                    
                    for r_idx, cell in enumerate(col_cells[1:], 2):
                        is_empty_row = all(ws.cell(r_idx, c).value in [None, ""] for c in range(1, ws.max_column + 1))
                        if is_empty_row: continue
                        r_type = ws.cell(r_idx, type_idx).value if type_idx else None
                        if cell.value is None and r_type not in ["MEAN", "MAX", "MIN", "STD", "W-AVG", "W-MAX", "W-MIN", "W-STD", "NEXT QUERY"]: continue
                        cell.border = thin_border
                        cell.fill = col_fill
                        if r_type == "MEAN": cell.font = Font(bold=True, color="1F4E78")
                        elif r_type == "MAX": cell.font = Font(italic=True, color="833C0C")
                        elif r_type == "MIN": cell.font = Font(italic=True, color="833C0C")
                        elif r_type == "STD": cell.font = Font(italic=True, color="9C5700")
                        elif r_type == "NEXT QUERY":
                            cell.font = Font(bold=True, color="FF0000")
                            cell.alignment = Alignment(horizontal="left")
                        elif r_type == "W-AVG": cell.font = Font(bold=True, color="203764")
                        elif r_type == "W-MAX": cell.font = Font(italic=True, color="542914")
                        elif r_type == "W-MIN": cell.font = Font(italic=True, color="542914")
                        elif r_type == "W-STD": cell.font = Font(italic=True, color="542914")
                        
                        if isinstance(cell.value, (int, float)):
                            if "diff%" in header: cell.number_format = '0.00%'
                            elif "speedup" in header: cell.number_format = '0.00"x"'
                            elif "time" in header:
                                val = abs(cell.value)
                                if val == 0: cell.number_format = '0.00'
                                elif val >= 100: cell.number_format = '0.00'
                                else:
                                    try:
                                        needed_decimals = max(2, math.ceil(-math.log10(val)) + 2)
                                        cell.number_format = '0.' + ('0' * needed_decimals)
                                    except: cell.number_format = '0.000'
                            elif "hpfr" in header: cell.number_format = '0.000'
            wb.save(self.filename)
        except Exception as e:
            print(f"Warning: Styling step failed: {e}")